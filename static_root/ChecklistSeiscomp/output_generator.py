import openpyxl

filename = 'template.xlsx'
wb = openpyxl.load_workbook(filename=filename)
ws = wb.worksheets[0]

# input:
metadata = {'kelompok': 2,
        'operator1': 'Fajar',
        'operator2': 'Tri',
        'tanggal': 'Minggu, 19 Februari 2023'}
gaps = ['ACJM', 'PCJI', 'BBLSI', 'BUMSI']
spikes = ['ACJM', 'PCJI', 'BBLSI', 'BUMSI'
]
blanks = ['ACJM', 'PCJI', 'BBLSI', 'BUMSI']

# constants
nongaransi1 = [ws.cell(row=i,column=2).value for i in range(8, 235 +1)]
nongaransi2 = [ws.cell(row=i,column=13).value for i in range(8, 181 +1)]
garansi = [ws.cell(row=i,column=13).value for i in range(187, 196 +1)]


def loop_cells(type, station_code, row, column):
    for i in range(len(station_code)):
        for j in range(len(type)):
            if type[j] == station_code[i]:
                ws.cell(row=i + row, column=column).value = 1


def clear_cells(cell_range: str, ws):
    for row in ws[cell_range]:
        for cell in row:
            cell.value = None


def clear_contents():
    clear_cells('E8:J235', ws)
    clear_cells('W8:AB181', ws)
    clear_cells('W187:AB196', ws)

def update_metadata(metadata, operator):
    ws['A3'] = f"KELOMPOK : {metadata['kelompok']}"

    if operator == 1:
        ws['L216'] = metadata['operator1']
        ws['AB3'] = metadata['tanggal']
    else:
        ws['R216'] = metadata['operator2']
        ws['AB3'] = metadata['tanggal']

def update_jam_12(gaps, spikes, blanks):
    clear_cells('E8:G235', ws)
    clear_cells('W8:Y181', ws)
    clear_cells('W187:Y196', ws)

    # nongaransi1 loop
    loop_cells(gaps, nongaransi1, 8, 5)
    loop_cells(spikes, nongaransi1, 8, 6)
    loop_cells(blanks, nongaransi1, 8, 7)

    # nongaransi2 loop
    loop_cells(gaps, nongaransi2, 8, 23)
    loop_cells(spikes, nongaransi2, 8, 24)
    loop_cells(blanks, nongaransi2, 8, 25)

    # garansi loop
    loop_cells(gaps, garansi, 187, 23)
    loop_cells(spikes, garansi, 187, 24)
    loop_cells(blanks, garansi, 187, 25)

def update_jam_00(gaps, spikes, blanks):
    clear_cells('H8:J235', ws)
    clear_cells('Z8:AB181', ws)
    clear_cells('Z187:AB196', ws)

    # nongaransi1 loop
    loop_cells(gaps, nongaransi1, 8, 8)
    loop_cells(spikes, nongaransi1, 8, 9)
    loop_cells(blanks, nongaransi1, 8, 10)

    # nongaransi2 loop
    loop_cells(gaps, nongaransi2, 8, 26)
    loop_cells(spikes, nongaransi2, 8, 27)
    loop_cells(blanks, nongaransi2, 8, 28)

    # garansi loop
    loop_cells(gaps, garansi, 187, 26)
    loop_cells(spikes, garansi, 187, 27)
    loop_cells(blanks, garansi, 187, 28)
    
update_jam_12(gaps, spikes, blanks)
update_jam_00(gaps, spikes, blanks)

update_metadata(metadata, 1)

wb.save(filename=filename)
