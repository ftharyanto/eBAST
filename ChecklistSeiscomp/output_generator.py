import openpyxl

def generate_excel(filename, response, metadata, data):
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb.worksheets[0]

    # constants
    nongaransi1 = [ws.cell(row=i,column=2).value for i in range(8, 267 +1)]
    nongaransi2 = [ws.cell(row=i,column=13).value for i in range(8, 149 +1)]
    garansi = [ws.cell(row=i,column=13).value for i in range(154, 246 +1)]


    def loop_cells(type, station_code, row, column):
        for i in range(len(station_code)):
            for j in range(len(type)):
                if type[j] == station_code[i]:
                    ws.cell(row=i + row, column=column).value = 1


    def clear_cells(cell_range: str, ws):
        for row in ws[cell_range]:
            for cell in row:
                cell.value = None


    def clear_contents():
        clear_cells('E8:J267', ws)
        clear_cells('T8:V149', ws)
        clear_cells('T154:V246', ws)

    def update_metadata(metadata):
        ws['A3'] = f"KELOMPOK : {metadata['kelompok']}"

        ws['I264'] = metadata['operator']
        ws['V3'] = metadata['tanggal']
        ws['A2'] = metadata['shift']

    def update_data(gaps, spikes, blanks):
        clear_cells('E8:G267', ws)
        clear_cells('T8:V149', ws)
        clear_cells('T154:V246', ws)

        # nongaransi1 loop
        loop_cells(gaps, nongaransi1, 8, 5)
        loop_cells(spikes, nongaransi1, 8, 6)
        loop_cells(blanks, nongaransi1, 8, 7)

        # nongaransi2 loop
        loop_cells(gaps, nongaransi2, 8, 15)
        loop_cells(spikes, nongaransi2, 8, 16)
        loop_cells(blanks, nongaransi2, 8, 17)

        # garansi loop
        loop_cells(gaps, garansi, 154, 15)
        loop_cells(spikes, garansi, 154, 16)
        loop_cells(blanks, garansi, 154, 17)
        
    update_data(data['gaps'], data['spikes'], data['blanks'])
    update_metadata(metadata)

    wb.save(response)